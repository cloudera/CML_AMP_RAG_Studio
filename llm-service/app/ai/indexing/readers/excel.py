#
#  CLOUDERA APPLIED MACHINE LEARNING PROTOTYPE (AMP)
#  (C) Cloudera, Inc. 2024
#  All rights reserved.
#
#  Applicable Open Source License: Apache 2.0
#
#  NOTE: Cloudera open source products are modular software products
#  made up of hundreds of individual components, each of which was
#  individually copyrighted.  Each Cloudera open source product is a
#  collective work under U.S. Copyright Law. Your license to use the
#  collective work is as provided in your written agreement with
#  Cloudera.  Used apart from the collective work, this file is
#  licensed for your use pursuant to the open source license
#  identified above.
#
#  This code is provided to you pursuant a written agreement with
#  (i) Cloudera, Inc. or (ii) a third-party authorized to distribute
#  this code. If you do not have a written agreement with Cloudera nor
#  with an authorized and properly licensed third party, you do not
#  have any rights to access nor to use this code.
#
#  Absent a written agreement with Cloudera, Inc. ("Cloudera") to the
#  contrary, A) CLOUDERA PROVIDES THIS CODE TO YOU WITHOUT WARRANTIES OF ANY
#  KIND; (B) CLOUDERA DISCLAIMS ANY AND ALL EXPRESS AND IMPLIED
#  WARRANTIES WITH RESPECT TO THIS CODE, INCLUDING BUT NOT LIMITED TO
#  IMPLIED WARRANTIES OF TITLE, NON-INFRINGEMENT, MERCHANTABILITY AND
#  FITNESS FOR A PARTICULAR PURPOSE; (C) CLOUDERA IS NOT LIABLE TO YOU,
#  AND WILL NOT DEFEND, INDEMNIFY, NOR HOLD YOU HARMLESS FOR ANY CLAIMS
#  ARISING FROM OR RELATED TO THE CODE; AND (D)WITH RESPECT TO YOUR EXERCISE
#  OF ANY RIGHTS GRANTED TO YOU FOR THE CODE, CLOUDERA IS NOT LIABLE FOR ANY
#  DIRECT, INDIRECT, INCIDENTAL, SPECIAL, EXEMPLARY, PUNITIVE OR
#  CONSEQUENTIAL DAMAGES INCLUDING, BUT NOT LIMITED TO, DAMAGES
#  RELATED TO LOST REVENUE, LOST PROFITS, LOSS OF INCOME, LOSS OF
#  BUSINESS ADVANTAGE OR UNAVAILABILITY, OR LOSS OR CORRUPTION OF
#  DATA.

import json
import logging
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Dict, Iterator, List, Optional, Sequence, Tuple

import pandas as pd
from openpyxl import load_workbook
from pyxlsb import open_workbook as open_xlsb_workbook
from llama_index.core.node_parser.interface import MetadataAwareTextSplitter
from llama_index.core.schema import Document, TextNode

from .base_reader import BaseReader, ChunksResult
from ....exceptions import DocumentParseError

logger = logging.getLogger(__name__)


class _ExcelSplitter(MetadataAwareTextSplitter):
    """Custom splitter for Excel files that handles multiple sheets and converts rows to JSON."""

    def split_text_metadata_aware(self, text: str, metadata_str: str) -> List[str]:
        # metadata_str is kept as an argument to satisfy the interface, but it is not used
        # because metadata is added to the chunks later.
        return self.split_text(text)

    def split_text(self, text: str) -> List[str]:
        """
        Expects text to be a JSON representation of the workbook:
        {"sheets": [{"name": "Sheet1", "rows": [...]}, ...]}
        Returns one JSON string per row with embedded sheet metadata.
        """
        try:
            workbook_data = json.loads(text)
            row_chunks: List[str] = []

            for sheet_data in workbook_data.get("sheets", []):
                sheet_name = sheet_data.get("name", "")
                rows = sheet_data.get("rows", [])

                for i, row in enumerate(rows):
                    # Embed sheet metadata in the row JSON
                    row_with_meta = {
                        "__sheet_name__": sheet_name,
                        "__row_number__": i + 1,
                        **row,
                    }
                    row_chunks.append(json.dumps(row_with_meta, sort_keys=True))

            return row_chunks
        except Exception as e:
            logger.error("Error splitting Excel text: %s", e)
            raise DocumentParseError("Error splitting Excel text") from e


class ExcelReader(BaseReader):
    """Reader that creates a Document from Excel files and uses splitter for chunking."""

    def load_chunks(self, file_path: Path) -> ChunksResult:
        ret = ChunksResult()

        try:
            # Build workbook data structure by streaming rows
            sheets_data: List[Dict[str, Any]] = []
            for sheet_name, row_number, row_dict in self._iter_rows(file_path):
                # Find or create sheet in sheets_data
                sheet_entry = next(
                    (s for s in sheets_data if s["name"] == sheet_name),
                    None,
                )
                if sheet_entry is None:
                    sheet_entry = {"name": sheet_name, "rows": []}
                    sheets_data.append(sheet_entry)
                sheet_entry["rows"].append(row_dict)

            if not sheets_data or all(not s["rows"] for s in sheets_data):
                logger.warning("No data rows found in Excel file %s", file_path)
                return ret

            # Create JSON representation of the workbook
            workbook_data = {"sheets": sheets_data}
            content = json.dumps(workbook_data)

            # Check for secrets in the serialized content
            secrets = self._block_secrets([content])
            if secrets is not None:
                ret.secret_types = secrets
                return ret

            # Optionally anonymize PII
            anonymized_text = self._anonymize_pii(content)
            if anonymized_text is not None:
                ret.pii_found = True
                content = anonymized_text

            # Create document and use splitter (following CSV reader pattern)
            document = Document(text=content)
            document.id_ = self.document_id
            self._add_document_metadata(document, file_path)

            local_splitter = _ExcelSplitter()

            try:
                rows = local_splitter.get_nodes_from_documents([document])
            except Exception as e:
                logger.error("Error processing Excel file %s: %s", file_path, e)
                return ret

            # Extract embedded metadata and clean up the row JSON
            for i, row in enumerate(rows):
                try:
                    assert isinstance(row, TextNode)
                    row_data = json.loads(row.text)
                    sheet_name = row_data.pop("__sheet_name__", "")
                    row_number = row_data.pop("__row_number__", i + 1)

                    # Store cleaned row JSON without metadata fields
                    row.text = json.dumps(row_data, sort_keys=True)

                    # Add metadata
                    row.metadata["file_name"] = document.metadata["file_name"]
                    row.metadata["document_id"] = document.metadata["document_id"]
                    row.metadata["data_source_id"] = document.metadata["data_source_id"]
                    row.metadata["chunk_number"] = i
                    row.metadata["row_number"] = row_number
                    row.metadata["sheet_name"] = sheet_name
                    row.metadata["chunk_format"] = "json"
                except Exception as e:
                    logger.error("Error processing row %d: %s", i, e)

            ret.chunks = rows
            return ret

        except Exception as exc:
            logger.error("Error reading Excel file %s: %s", file_path, exc)
            return ret

    def _iter_rows(self, file_path: Path) -> Iterator[Tuple[str, int, Dict[str, str]]]:
        suffix = file_path.suffix.lower()

        if suffix == ".xlsb":
            try:
                yield from self._iter_xlsb_rows(file_path)
                return
            except Exception as exc:
                logger.warning(
                    "pyxlsb streaming failed for %s: %s. Falling back to pandas.",
                    file_path,
                    exc,
                )

        if suffix == ".xls":
            try:
                yield from self._iter_openpyxl_rows(file_path)
                return
            except Exception as exc:
                logger.warning(
                    "openpyxl streaming failed for %s: %s. Falling back to pandas.",
                    file_path,
                    exc,
                )

        yield from self._iter_pandas_rows(file_path)

    def _iter_openpyxl_rows(
        self, file_path: Path
    ) -> Iterator[Tuple[str, int, Dict[str, str]]]:
        workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
        try:
            for worksheet in workbook.worksheets:
                headers: Optional[List[str]] = None

                for row_index, row in enumerate(
                    worksheet.iter_rows(values_only=True), start=1
                ):
                    values = [self._stringify_value(cell) for cell in row]

                    if headers is None:
                        headers = self._normalize_headers(values)
                        continue

                    if not headers:
                        continue

                    self._ensure_header_length(headers, len(values))
                    row_dict = self._build_row_dict(headers, values)

                    if self._is_empty_row(list(row_dict.values())):
                        continue

                    yield worksheet.title, row_index, row_dict
        finally:
            workbook.close()

    def _iter_xlsb_rows(
        self, file_path: Path
    ) -> Iterator[Tuple[str, int, Dict[str, str]]]:
        with open_xlsb_workbook(str(file_path)) as workbook:
            for sheet_name in workbook.sheets:
                with workbook.get_sheet(sheet_name) as sheet:
                    headers: Optional[List[str]] = None

                    for row_index, row in enumerate(sheet.rows(), start=1):
                        values = [self._stringify_value(cell.v) for cell in row]

                        if headers is None:
                            headers = self._normalize_headers(values)
                            continue

                        if not headers:
                            continue

                        self._ensure_header_length(headers, len(values))
                        row_dict = self._build_row_dict(headers, values)

                        if self._is_empty_row(list(row_dict.values())):
                            continue

                        yield sheet_name, row_index, row_dict

    def _iter_pandas_rows(
        self, file_path: Path
    ) -> Iterator[Tuple[str, int, Dict[str, str]]]:
        try:
            sheets = pd.read_excel(file_path, sheet_name=None, engine="calamine")
        except Exception as exc:
            logger.error("Error reading Excel file %s: %s", file_path, exc)
            return

        if len(sheets) == 0:
            logger.error("No sheets found in Excel file %s", file_path)
            return

        for sheet_name, df in sheets.items():
            if df.empty:
                continue

            df = df.map(str)
            headers = [str(column) for column in df.columns]
            row_counter = 1  # headers come from the first row in pandas

            for record in df.itertuples(index=False, name=None):
                row_counter += 1
                values = [str(value) if value is not None else "" for value in record]
                self._ensure_header_length(headers, len(values))
                row_dict = self._build_row_dict(headers, values)

                if self._is_empty_row(list(row_dict.values())):
                    continue

                yield str(sheet_name), row_counter, row_dict

    @staticmethod
    def _normalize_headers(values: Sequence[Any]) -> List[str]:
        headers: List[str] = []
        seen: Dict[str, int] = {}

        for index, value in enumerate(values):
            base = str(value).strip() if value not in (None, "") else ""
            if not base:
                base = f"column_{index + 1}"

            count = seen.get(base, 0)
            header = f"{base}_{count}" if count else base
            seen[base] = count + 1
            headers.append(header)

        if not headers:
            headers.append("column_1")

        return headers

    @staticmethod
    def _ensure_header_length(headers: List[str], value_count: int) -> None:
        while len(headers) < value_count:
            headers.append(f"column_{len(headers) + 1}")

    @staticmethod
    def _build_row_dict(
        headers: Sequence[str], values: Sequence[str]
    ) -> Dict[str, str]:
        row: Dict[str, str] = {}
        for index, header in enumerate(headers):
            row[header] = values[index] if index < len(values) else ""
        return row

    @staticmethod
    def _is_empty_row(values: Sequence[str]) -> bool:
        return all(value == "" for value in values)

    @staticmethod
    def _stringify_value(value: Any) -> str:
        if value is None:
            return ""

        if isinstance(value, (datetime, date)):
            return value.isoformat()

        if isinstance(value, time):
            try:
                return value.isoformat()
            except ValueError:
                return value.strftime("%H:%M:%S")

        return str(value)
